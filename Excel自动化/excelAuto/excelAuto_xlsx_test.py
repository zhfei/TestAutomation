#!~/Documents/pythonFiles/venv python
# -*- coding:utf-8 -*-

# xls 与 xlsx的读写方式不同


# 1、使用xlrd模块对xls文件进行读操作
# 2、使用xlwt模块对xls文件进行写操作
# 3、使用openpyxl模块对xlsx文件进行读操作
# 4、使用openpyxl模块对xlsx文件进行写操作
# 5、修改已经存在的工作簿(表)

# xlsx读写
import openpyxl

#网络请求
import requests
import json



# 读取测试用例
excelDir = u'/Users/zhoufei/Documents/接口自动化/Excel自动化/excel自动化测试.xlsx'
excel_res = 'pass'


#-----------------读xlsx-------------------

# 1.获取工作簿
workbook_xlsx = openpyxl.load_workbook(excelDir)

# 2.获取所有表名
sheetnames_xlsx = workbook_xlsx.sheetnames
print(sheetnames_xlsx)

# 3.获取工作表对象(根据名字)
worksheet_xlsx = workbook_xlsx[sheetnames_xlsx[0]]
print(worksheet_xlsx)
# 3.获取工作表对象(根据下标)
worksheet_xlsx = workbook_xlsx.worksheets[0]
print(worksheet_xlsx)
# 3.获取工作表对象(根据当前活跃的sheet)
worksheet_xlsx = workbook_xlsx.active
print(worksheet_xlsx)

# 4.获取工作表属性
title_xlsx = worksheet_xlsx.title
print(title_xlsx)
rows_xlsx = worksheet_xlsx.max_row
print(rows_xlsx)
columns_xlsx = worksheet_xlsx.max_column
print(columns_xlsx)

# 5.按照 行 获取表中的数据
for row in worksheet_xlsx.rows:
    print(row)
    for cell in row:
        print(cell.value, end='')
    print('\n')
# # 读第3行
# for row in list(worksheet_xlsx.rows)[1]:
#     print(row)
#     for cell in row:
#         print(cell.value, end='')
#     print('\n')
#
# # 读切片
# for row in list(worksheet_xlsx.rows)[0:3]:
#     print(row)
#     for cell in row[0:3]:
#         print(cell.value, end='')
#     print('\n')

content_cell = worksheet_xlsx.cell(row=1, column=1).value
print(content_cell)

print("-" * 20)

# 5.按照 列 获取表中的数据
for column in worksheet_xlsx.columns:
    for cell in column:
        print(cell.value, end='')
    print('\n')



#-----------------写xlsx-------------------

# 1.创建workbook和worksheet
#创建一个workbook对象（即：创建一个Excel文件）
workbook = openpyxl.Workbook()
#获取当前活跃的worksheet（默认第一个）
worksheet = workbook.active
worksheet.title = "python 创建的excel的sheet"

# worksheet2 = workbook.create_sheet(1) #(默认插入工作薄第一个位置)
worksheet2 = workbook.create_sheet() #(默认插入工作薄尾部)

# 2.写数据到表中
Province = ['北京市', '天津市', '河北省', '山西省', '内蒙古自治区', '辽宁省',
            '吉林省', '黑龙江省', '上海市', '江苏省', '浙江省', '安徽省', '福建省',
            '江西省', '山东省', '河南省', '湖北省', '湖南省', '广东省', '广西壮族自治区',
            '海南省', '重庆市', '四川省', '贵州省', '云南省', '西藏自治区', '陕西省', '甘肃省',
            '青海省', '宁夏回族自治区', '新疆维吾尔自治区']

Income = ['5047.4', '3247.9', '1514.7', '1374.3', '590.7', '1499.5', '605.1', '654.9',
          '6686.0', '3104.8', '3575.1', '1184.1', '1855.5', '1441.3', '1671.5', '1022.7',
          '1199.2', '1449.6', '2906.2', '972.3', '555.7', '1309.9', '1219.5', '715.5', '441.8',
          '568.4', '848.3', '637.4', '653.3', '823.1', '254.1']

Project = ['各省市', '工资性收入', '家庭经营纯收入', '财产性收入', '转移性收入', '食品', '衣着',
           '居住', '家庭设备及服务', '交通和通讯', '文教、娱乐用品及服务', '医疗保健', '其他商品及服务']

#第一行放Project，类型
for i in range(len(Project)):
    worksheet.cell(1,i+1,Project[i])

#第一列放Province,省份
for i in range(len(Province)):
    worksheet.cell(i+2,1,Province[i])

#第二行放Income,收入
for i in range(len(Income)):
    worksheet.cell(2,i+2,Income[i])

workbook.save(u'/Users/zhoufei/Documents/接口自动化/Excel自动化/excel自动化测试_xlsx_res.xlsx')




#-----------------excel测试用例结果写入-------------------

# 测试结果写入excel (xlsx方式)

excel_res = 'pass'
workbook_r = openpyxl.load_workbook(excelDir)
sheet_r = workbook_r.worksheets[0]
sheet_r.cell(2,9, excel_res)
workbook_r.save('/Users/zhoufei/Documents/接口自动化/Excel自动化/excel自动化测试_res.xlsx')








